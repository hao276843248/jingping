import json
from datetime import datetime, date
import sys
import formulas
import openpyxl
import openpyxl.cell.cell
import pandas as pd

from util import ret_row_column, get_row_column, get_list_row, I_A_Z_char_dict


def get_value(value, sheet):
    STATE_DATE = date(1900, 1, 1).toordinal() - 2
    row, column = ret_row_column(value.upper())
    va = sheet.cell(row, column).value
    if isinstance(va, datetime):
        return datetime.date(va).toordinal() - STATE_DATE
    return va


def get_db2_data():
    # 获取db2中的数据
    with open(sys.path[0] + "/db2_data.txt", "r", encoding="utf-8") as f:
        data = f.read()
        data = eval(data)
        return data


def get_date():
    print(sys.path)
    # 获取总表数据
    with open(sys.path[0] + "/excel_data_json.db", "r", encoding="utf-8") as f:
        data = f.read()
        data = json.loads(data)
    return data


def change_jiashe():
    """
    修改假设表 date
    :return:
    """
    a = openpyxl.load_workbook('VBA.xlsm', data_only=False)
    # a = openpyxl.load_workbook('data.xlsx', data_only=False)
    # # set_data(a)
    sheet = a.worksheets[4]
    # # print(type(get_value("E4", sheet)))
    data["假设表"]["E" + str(1)] = "J"
    for i in range(2, 378):
        value = get_value("E" + str(i), sheet)
        if isinstance(value, str) and value.startswith("="):
            print(value)
            func = formulas.Parser().ast(value)[1].compile()
            rows = {}
            for j in func.inputs:
                col, row = get_row_column(j)
                if str(row) in rows:
                    rows[str(row)] += 1
                else:
                    rows[str(row)] = 1
            str_row = "J" + sorted(rows, key=lambda k: rows[k], reverse=True)[0]
            value = value.replace("INDIRECT($E$1&ROW())", str_row)
            print(value)
        data["假设表"]["E" + str(i)] = value
        # data["财报"]["E" + str(i)] =
    with open("data_假设表修改后的_json.txt", "w", encoding="utf-8") as f:
        f.write(json.dumps(data))


def get_date_value(value: str, sheet: str):
    global data
    """
    获取data中value的值
    :param value: 位置  A1
    :param sheet: 表名  "参数"
    :return: str
    """
    return data[sheet].setdefault(value, None)


def set_data_caibao():
    a = openpyxl.load_workbook('VBA.xlsm', data_only=False)
    # a = openpyxl.load_workbook('data.xlsx', data_only=False)
    # # set_data(a)
    sheet = a.worksheets[5]
    for j in get_list_row("NC"):
        for i in range(1, 691):
            data["财报"][j + str(i)] = get_value(j + str(i), sheet)
    with open("data_财报表不用修改了_json.txt", "w", encoding="utf-8") as f:
        f.write(json.dumps(data))


def set_data_qitaxiangmuxinin():
    a = openpyxl.load_workbook('VBA.xlsm', data_only=False)
    # a = openpyxl.load_workbook('data.xlsx', data_only=False)
    # # set_data(a)
    sheet = a.worksheets[9]
    for j in get_list_row("E"):
        for i in range(1, 10):
            data["其他项目信息"][j + str(i)] = get_value(j + str(i), sheet)
    with open("data_财报表修改后的_json.txt", "w", encoding="utf-8") as f:
        f.write(json.dumps(data))


def set_data(a):
    # 读取excel中所有单元格内容
    lists = [[2, "M", 257], [3, "X", 67], [4, "U", 378], [5, "NC", 691],
             [6, "AC", 307], [7, "M", 451], [8, "AS", 132], [9, "E", 10], [10, "E", 38], [11, "E", 12],
             [12, "E", 51], [13, "E", 28], [14, "E", 18], [15, "Y", 36], [16, "AK", 223], [17, "H", 50], [18, "S", 40]]
    for k in lists:
        # 第几个工作表
        sheet = a.worksheets[k[0]]
        # 获得sheet表名称
        print("Worksheet name(s):", sheet.title)
        # # print(dir(sheet))
        data = {}
        with open("data.txt", "r", encoding="utf-8") as f:
            data = f.read()
            data = eval(data)
        data[sheet.title] = {}
        lie_list = get_list_row(k[1])
        for i in range(1, k[2]):
            for j in lie_list:
                data[sheet.title][j + str(i)] = get_value(j + str(i), sheet)
        with open("data.txt", "w", encoding="utf-8") as f:
            f.write(str(data))


def set_bug_data():
    data["假设表"]["E256"] = "=E219*(1+敏感性!$D$4)"
    data["假设表"]["E257"] = "=E220*(1+敏感性!$D$4)"
    data["假设表"]["E258"] = "=E221*(1+敏感性!$D$4)"
    for i in range(1, 378):
        if data["假设表"]['H' + str(i)] == None:
            data["假设表"]['H' + str(i)] = ""
        if data["假设表"]['J' + str(i)] == None:
            data["假设表"]['J' + str(i)] = ""
        if data["假设表"]['G' + str(i)] == None:
            data["假设表"]['G' + str(i)] = ""
    with open("excel_data_json.db", "w", encoding="utf-8") as f:
        f.write(json.dumps(data))


def set_T9_to_J66():
    data['参数']['T9'] = "=J66"
    with open("excel_data_json.db", "w", encoding="utf-8") as f:
        f.write(json.dumps(data))


def set_fan_type_data(turbine, number=1):
    global data
    db2 = get_db2_data()
    df = pd.DataFrame(db2)
    val = df.loc[df[1] == str(turbine)]
    if not val.loc[::][0].tolist():
        val = pd.DataFrame([[0] * 28])
    list_come = get_list_row("AG")
    dicts_db2 = {1: "56", 2: "57", 3: "58"}
    # 机型高度
    dicts_gaodu = {1: "58", 2: "59", 3: "60"}
    # 机型价格
    dicts_jaige = {1: "66", 2: "67", 3: "68"}
    # 钢塔-重量
    dicts_gangzhong = {1: "72", 2: '77', 3: '82'}
    # 锚栓-重量
    dicts_maozhong = {1: "73", 2: '78', 3: '83'}
    # 混塔段-总价
    dicts_hunta = {1: "73", 2: '78', 3: '83'}
    # 整机吊安装费（单价）
    dicts_diaozhuang = {1: "75", 2: '80', 3: '85'}
    # 垫层混凝土
    dicts_dchnt = {1: "88", 2: '93', 3: '98'}
    # 混凝土
    dicts_hnt = {1: "89", 2: '94', 3: '99'}
    # 钢筋
    dicts_gj = {1: "90", 2: '95', 3: '100'}

    for i, v in enumerate(val.iloc[0]):
        if i == 1:
            continue
        # print(list_come[i + 1] + dicts_db2[number])
        # print(v)
        # print(type(v))
        data["DB2"][list_come[i + 1] + dicts_db2[number]] = v

    data["模型"]["E" + dicts_gaodu[number]] = val.iloc[0][9]
    # data["模型"]["E" + dicts_jaige[number]] = val.iloc[0][9]
    data["模型"]["E" + dicts_gangzhong[number]] = val.iloc[0][11]
    data["模型"]["E" + dicts_maozhong[number]] = val.iloc[0][22]
    data["模型"]["E" + dicts_diaozhuang[number]] = val.iloc[0][26]
    data["模型"]["E" + dicts_dchnt[number]] = val.iloc[0][16]
    data["模型"]["E" + dicts_hnt[number]] = val.iloc[0][17]
    data["模型"]["E" + dicts_gj[number]] = val.iloc[0][18]


def check_data_to_list(data):
    # data = {}
    data_list = {1: [[]]}
    b = [[2, "M", 257, "模型"], [3, "X", 67, "参数"], [4, "U", 378, "假设表"], [5, "NC", 691, "财报"],
         [6, "AC", 307, "DB2"], [7, "M", 451, "概算表"], [8, "AS", 132, "输出"], [9, "E", 10, "其他项目信息"],
         [10, "E", 38, "其他市场信息"], [11, "E", 12, "其他融资信息"],
         [12, "E", 51, "其他运维信息"], [13, "E", 28, "其他价格信息"], [14, "E", 18, "其他工程信息"], [15, "Y", 36, "敏感性"],
         [16, "AK", 223, "可研"], [17, "H", 50, "DB1"], [18, "S", 40, "DB2"]]
    for ib in b:
        for k1, v1 in data[ib[3]].items():
            _, max_column = ret_row_column(str(ib[1]) + "1")
            max_row = ib[2]
            # a     1
            row, column = ret_row_column(k1)
            if ib[3] not in data_list.keys():
                lista = []
                for i in range(max_column+1):
                    lista.append([None] * (max_row+1))
                data_list[ib[3]] = lista
            try:
                data_list[ib[3]][column][row] = v1
            except:
                print(column)
                print(row)
    return data_list


if __name__ == '__main__':
    """
    记录一下AL:JQ 16,12 =1 
    记录一下T:AK 15,11 =1 
    记录一下H:S 14,10 =1 
    """
    data = get_date()
    vb = check_data_to_list(data)
    with open("data_list_json.db", "w", encoding="utf-8") as f:
        f.write(json.dumps(vb))
    # print(sdata_list[0][0])
    # print(type(sdata_list))
